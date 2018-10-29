from numpy import *
import numpy as np
import sympy
import openpyxl
from openpyxl import Workbook
'''Input Parameter'''
N=30000
a=0;b=300
epsilon2=0.3
Z=1.0
mu=1.0
dx=(b-a)/N
'''Functions'''
def sech(x):
	return(sympy.cosh(x)**(-1))
Ni=[]
def ni(x):
	u0=1.0
	x0=100
	epsilon=double(epsilon2**(0.5))
	return(double(1+3*epsilon**(2)*u0*sech(epsilon*u0**(0.5)*(x-x0)/2)**(2)))
for i in range(N+1):
	Ni.append(ni((i+1)*dx))
x=[]
for i in range(N+1):
	x.append(i*dx)
wb = Workbook()
ws = wb.active
i=0
while i<N+1:
	ws.cell(row=i+1,column=1).value=x[i]
	ws.cell(row=i+1,column=2).value=Ni[i]
	i=i+1
wb.save('result.xlsx')

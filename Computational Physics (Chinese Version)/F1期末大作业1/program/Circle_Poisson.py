import numpy as np
import openpyxl
from openpyxl import Workbook
wb=Workbook()
wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
#`\textcolor{darkgreen}{常数设置}`
pi=3.1415926535897932384626433
#`\textcolor{darkgreen}{参数设置}`
ba=-5.0;bb=5.0;bc=5.0;bd=-5.0
nmin=5;nmax=18;x0=2.0;y0=0.0;rho=1
for index in range(nmin,nmax):
	M=index-1;N=1+(index-1)*(index-1);Deltar=bb/index;Deltatheta=2*pi/index
	#`\textcolor{darkgreen}{生成系数初始化}`
	a=np.zeros(shape=M+1)
	b=np.zeros(shape=M+1)
	c=np.zeros(shape=M+1)
	d=np.zeros(shape=M+1)
	e=np.zeros(shape=M+1)
	for i in range(1,M+1):
		a[i]=2.0/(Deltar*Deltar)+2.0/(i*Deltar*i*Deltar*Deltatheta*Deltatheta)
		b[i]=-1.0/(i*Deltar*i*Deltar*Deltatheta*Deltatheta)
		c[i]=-1.0/(i*Deltar*i*Deltar*Deltatheta*Deltatheta)
		d[i]=-1.0*(i-0.5)*Deltar/(i*Deltar*Deltar*Deltar)
		e[i]=-1.0*(i+0.5)*Deltar/(i*Deltar*Deltar*Deltar)
	m=(2.0*M*Deltatheta)/(pi*Deltar*Deltar)
	n=-1.0*(2.0*Deltatheta)/(pi*Deltar*Deltar)
	#`\textcolor{darkgreen}{数组矩阵初始化}`
	A=np.zeros(shape=(N,N))
	A[0][0]=m
	for i in range(1,M+1):
		A[0][i]=n
	for i in range(1,M+1):
		A[i][0]=d[1]
	for i in range(N):
		if((i!=0)and(i+M-N)<0):
			A[i][(i+M)-N]=e[int((i-1)/M)-M]
	for i in range(N):
		if((i!=0)and(i+M-N)<0):
			A[(i+M)-N][i]=d[int((i-1)/M)+1-M]
	for k in range(M):
		B=np.zeros(shape=(M,M))
		for l in range(M):
			B[l-1][l]=c[k+1]
			B[l][l]=a[k+1]
			B[l][l-1]=b[k+1]
		for i in range(1+k*M,M+1+k*M):
			for j in range(1+k*M,M+1+k*M):
				A[i][j]=B[i-k*M-1][j-k*M-1]
	b=np.zeros(shape=(N))
	b[int(np.sqrt(x0*x0+y0*y0)/Deltar)+int(np.arctan(y0/x0)/Deltatheta)]=rho
#`\textcolor{darkgreen}{生成并储存矩阵}`
	sheet=wb.create_sheet("Sheet%d"%(index-nmin+1))
	for i in range(N):
		sheet.cell(row=i+1,column=N+1,value=b[i])
		for j in range(N):
			sheet.cell(row=i+1,column=j+1,value=A[i][j])
wb.save('Cmatrix.xlsx')

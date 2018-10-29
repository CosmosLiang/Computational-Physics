import numpy as np
import time
import method as med
import openpyxl
from openpyxl import Workbook
#`\textcolor{darkgreen}{直角坐标线性方程组输入}`
#wb=openpyxl.load_workbook('Rmatrix.xlsx')
#`\textcolor{darkgreen}{极坐标线性方程组输入}`
wb=openpyxl.load_workbook('Cmatrix.xlsx')
sheets=wb.get_sheet_names()
t1=np.zeros(shape=(len(sheets)))
t2=np.zeros(shape=(len(sheets)))
t3=np.zeros(shape=(len(sheets)))
t4=np.zeros(shape=(len(sheets)))
GridNumber=np.zeros(shape=(len(sheets)))
for index in range(1,len(sheets)+1):
	sheet=wb.get_sheet_by_name('Sheet'+str(index))
	A=np.zeros(shape=(sheet.max_row,sheet.max_row))
	b=np.zeros(shape=(sheet.max_row))
	GridNumber[index-1]=(sheet.max_row)
	for k in range(1,sheet.max_column):
		b[k-1]=sheet.cell(row=k,column=sheet.max_column).value
	for i in range(1,sheet.max_column):
		for j in range(1,sheet.max_column):
		    A[i-1][j-1]=sheet.cell(row=i,column=j).value
	time1=time.clock()
	med.gaussian_elimination(A,b)
	time2=time.clock()
	med.jacobi(A,b)
	time3=time.clock()
	med.gauss_seidel(A,b)
	time4=time.clock()
	med.sor(A,b)
	time5=time.clock()
	t1[index-1]=time2-time1
	t2[index-1]=time3-time2
	t3[index-1]=time4-time3
	t4[index-1]=time5-time4

	if index==5:#`\textcolor{darkgreen}{选择不同的index从而输出分格个数不同的解的图像}`
		import scipy.linalg
		import matplotlib.pyplot as plt
		from mpl_toolkits.mplot3d import Axes3D
		from scipy.interpolate import griddata
		from matplotlib import cm
		from matplotlib.ticker import LinearLocator,FormatStrFormatter

		#`\textcolor{darkgreen}{直角坐标设置}`
		#ba=-5.0;bb=5.0;bc=5.0;bd=-5.0;nmin=5;h=(bc-ba)/(index+nmin-1)
		#`\textcolor{darkgreen}{位置坐标设置}`
		#x=np.repeat(np.arange(ba+h,bc,h),index+nmin-2)
		#y=np.tile(np.arange(bd+h,bb,h),index+nmin-2)

		#`\textcolor{darkgreen}{极坐标设置}`
		ba=-5.0;bb=5.0;bc=5.0;bd=-5.0
		#`\textcolor{darkgreen}{常数设置}`
		pi=3.1415926535897932384626433
		#`\textcolor{darkgreen}{参数设置}`
		nmin=5;x0=1.0;y0=1.0;rho=1
		M=index-1;N=1+(index+nmin-2)*(index+nmin-2)
		Deltar=bb/(index+nmin-1);Deltatheta=2*pi/(index+nmin-1)
		#`\textcolor{darkgreen}{位置坐标设置}`
		x=np.zeros(shape=N)
		x[0]=0
		y=np.zeros(shape=N)
		y[0]=0
		for i in range(1,M+1):
			for j in range(1,M+1):
				x[i+j]=(i*Deltar)*np.cos(j*Deltatheta)
				y[i+j]=(i*Deltar)*np.sin(j*Deltatheta)

		#`\textcolor{darkgreen}{绘制解的图像部分}`
		phi1=med.gaussian_elimination(A,b)
		phi2=med.jacobi(A,b)
		phi3=med.gauss_seidel(A,b)
		phi4=med.sor(A,b)
		print(len(x),len(y),len(phi1))
		print(x,y,phi1)
		data=np.c_[x,y,phi1]
		#X,Y=np.meshgrid(np.arange(ba+h,bc,h),np.arange(bd+h,bb,h))
		X,Y=np.meshgrid(np.sort(x),np.sort(y))
		XX=X.flatten()
		YY=Y.flatten()
		#`\textcolor{darkgreen}{由于图像并不连续我们可以选择幂级数进行拟合，这里选择的是二次幂级数。}`
		A=np.c_[np.ones(data.shape[0]),data[:,:2],np.prod(data[:,:2],axis=1),
data[:,:2]**2]
		C,_,_,_=scipy.linalg.lstsq(A,data[:,2])
		Z = np.dot(np.c_[np.ones(XX.shape),XX,YY,XX*YY,XX**2,YY**2],C).reshape(X.shape)
		fig=plt.figure()
		ax=fig.gca(projection='3d')
		surf=ax.plot_surface(X,Y,Z,rstride=1,cstride=1,cmap='rainbow')
		ax.set_xlim(ba,bc)
		ax.set_ylim(bd,bb)
		ax.set_zlim(0,np.max(phi1))
		ax.zaxis.set_major_locator(LinearLocator(10))
		ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
		ax.set_xlabel('x')
		ax.set_ylabel('y')
		ax.set_zlabel('$\phi$')
		fig.colorbar(surf,shrink=0.5,aspect=5)
		#fname='figureR.eps'
		fname='figureC.eps'
		plt.savefig(fname)
		plt.clf()
	#`\textcolor{darkgreen}{为了避免覆盖节约内存我们在每次循环的时候将重复变量删除}`
	del sheet,A,b,time1,time2,time3,time4,time5

'''
#`\textcolor{darkgreen}{这部分代码可以实现将不同方法生成的图像分别绘制并储存}`
		for fn in range(1,5):
			varName='phi%d'%fn
			data=np.c_[x,y,locals()[varName]]
			fig=plt.figure()
			X,Y=np.meshgrid(np.arange(ba+h,bc,h),np.arange(bd+h,bb,h))
			XX=X.flatten()
			YY=Y.flatten()
			A=np.c_[np.ones(data.shape[0]),data[:,:2],np.prod(data[:,:2],axis=1),
data[:,:2]**2]
			C,_,_,_=scipy.linalg.lstsq(A,data[:,2])
			Z = np.dot(np.c_[np.ones(XX.shape),XX,YY,XX*YY,XX**2,YY**2],C)
.reshape(X.shape)
			fig=plt.figure()
			ax=fig.gca(projection='3d')
			surf=ax.plot_surface(X,Y,Z,rstride=1,cstride=1,cmap='rainbow')
			surf=ax.plot_surface(X,Y,Z,rstride=1,cstride=1,cmap='rainbow')
			ax.set_xlim(ba,bc)
			ax.set_ylim(bd,bb)
			ax.set_zlim(0,np.max(phi1))
			ax.zaxis.set_major_locator(LinearLocator(10))
			ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
			ax.set_xlabel('x')
			ax.set_ylabel('y')
			ax.set_zlabel('$\phi$')
			fig.colorbar(surf,shrink=0.5,aspect=5)
			fname='figR%d.eps'%fn
			#fname='figC%d.eps'%fn
			plt.savefig(fname)
			plt.clf()
			del data
'''

#`\textcolor{darkgreen}{绘制判断不同方法收敛速度快慢部分}`
import matplotlib.pyplot as plt
plt.figure()
plt.grid()
plt.plot(GridNumber,t1,'ro-',label="Gauss Elimination",color="red",linewidth=2)
plt.plot(GridNumber,t2,'ro-',label="Jacobi Method",color="orange",linewidth=2)
plt.plot(GridNumber,t3,'ro-',label="Gauss Sheidel",color="green",linewidth=2)
plt.plot(GridNumber,t4,'ro-',label="SOR Method",color="blue",linewidth=2)
plt.xlabel("Grid Number")
plt.ylabel("Time(ms)")
plt.xlim(0,GridNumber[-1]*1.1)
plt.ylim(0,t1[-1]*1.1)
plt.legend()
#plt.savefig("timeR.eps")
plt.savefig("timeC.eps")

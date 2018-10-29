import numpy as np
import time
import method as med
import openpyxl
from openpyxl import Workbook
#`\textcolor{darkgreen}{直角坐标线性方程组输入}`
wb=openpyxl.load_workbook('Rmatrix.xlsx')
#`\textcolor{darkgreen}{极坐标线性方程组输入}`
#wb=openpyxl.load_workbook('Cmatrix.xlsx')
sheets=wb.get_sheet_names()
t1=np.zeros(shape=(len(sheets)))
t2=np.zeros(shape=(len(sheets)))
t3=np.zeros(shape=(len(sheets)))
t4=np.zeros(shape=(len(sheets)))
GridNumber=np.zeros(shape=(len(sheets)))
for index in range(1,len(sheets)+1):
	sheet=wb.get_sheet_by_name('Sheet'+str(index))
	A=np.zeros(shape=(sheet.max_row,sheet.max_row))
	b=np.zeros(shape=(sheet.max_row))
	GridNumber[index-1]=(sheet.max_row)
	for k in range(1,sheet.max_column):
		b[k-1]=sheet.cell(row=k,column=sheet.max_column).value
	for i in range(1,sheet.max_column):
		for j in range(1,sheet.max_column):
		    A[i-1][j-1]=sheet.cell(row=i,column=j).value
	time1=time.clock()
	med.gaussian_elimination(A,b)
	time2=time.clock()
	med.jacobi(A,b)
	time3=time.clock()
	med.gauss_seidel(A,b)
	time4=time.clock()
	med.sor(A,b)
	time5=time.clock()
	t1[index-1]=time2-time1
	t2[index-1]=time3-time2
	t3[index-1]=time4-time3
	t4[index-1]=time5-time4
	#`\textcolor{darkgreen}{为了避免覆盖节约内存我们在每次循环的时候将重复变量删除}`
	del sheet,A,b,time1,time2,time3,time4,time5
#`\textcolor{darkgreen}{绘制判断不同方法收敛速度快慢部分}`
import matplotlib.pyplot as plt
plt.figure()
plt.grid()
plt.plot(GridNumber,t1,'ro-',label="Gauss Elimination",color="red",linewidth=2)
plt.plot(GridNumber,t2,'ro-',label="Jacobi Method",color="orange",linewidth=2)
plt.plot(GridNumber,t3,'ro-',label="Gauss Sheidel",color="green",linewidth=2)
plt.plot(GridNumber,t4,'ro-',label="SOR Method",color="blue",linewidth=2)
plt.xlabel("Grid Number")
plt.ylabel("Time(ms)")
plt.xlim(0,GridNumber[-1]*1.1)
plt.ylim(0,t1[-1]*1.1)
plt.legend()
plt.savefig("timeR.eps")
#plt.savefig("timeC.eps")

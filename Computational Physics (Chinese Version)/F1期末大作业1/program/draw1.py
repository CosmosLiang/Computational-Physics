import numpy as np
import scipy.linalg
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import griddata
from matplotlib import cm
from matplotlib.ticker import LinearLocator,FormatStrFormatter
import method as med
import openpyxl
from openpyxl import Workbook
#`\textcolor{darkgreen}{直角坐标线性方程组输入}`
#wb=openpyxl.load_workbook('Rmatrix.xlsx')
#`\textcolor{darkgreen}{极坐标线性方程组输入}`
wb=openpyxl.load_workbook('Cmatrix.xlsx')
sheets=wb.get_sheet_names()
GridNumber=np.zeros(shape=(len(sheets)))
#`\textcolor{darkgreen}{选择不同的index从而输出分格个数不同的解的图像}`
index=10#`\textcolor{darkgreen}{选择Sheet10进行计算}`
sheet=wb.get_sheet_by_name('Sheet'+str(index))
A=np.zeros(shape=(sheet.max_row,sheet.max_row))
b=np.zeros(shape=(sheet.max_row))
GridNumber[index-1]=(sheet.max_row)
for k in range(1,sheet.max_column):
	b[k-1]=sheet.cell(row=k,column=sheet.max_column).value
for i in range(1,sheet.max_column):
	for j in range(1,sheet.max_column):
		A[i-1][j-1]=sheet.cell(row=i,column=j).value
'''
#`\textcolor{darkgreen}{直角坐标设置}`
ba=-5.0;bb=5.0;bc=5.0;bd=-5.0;nmin=5;h=(bc-ba)/(index+nmin-1)
#`\textcolor{darkgreen}{位置坐标设置}`
xx=np.arange(ba+h,bc,h)
x=np.repeat(xx,index+nmin-2)
yy=np.arange(bd+h,bb,h)
y=np.tile(yy,index+nmin-2)
'''
#`\textcolor{darkgreen}{极坐标设置}`
ba=-5.0;bb=5.0;bc=5.0;bd=-5.0
#`\textcolor{darkgreen}{常数设置}`
pi=3.1415926535897932384626433
#`\textcolor{darkgreen}{参数设置}`
nmin=5;x0=1.0;y0=1.0;rho=1
M=index-1;N=1+(index+nmin-2)*(index+nmin-2)
Deltar=bb/(index+nmin-1);Deltatheta=2*pi/(index+nmin-1)
#`\textcolor{darkgreen}{位置坐标设置}`
x=np.zeros(shape=N);x[0]=0
y=np.zeros(shape=N);y[0]=0
for i in range(1,M+1):
	for j in range(1,M+1):
		x[i*j]=(i*Deltar)*np.cos(j*Deltatheta)
		y[i*j]=(i*Deltar)*np.sin(j*Deltatheta)

#`\textcolor{darkgreen}{绘制解的图像部分}`
phi1=med.gaussian_elimination(A,b)
phi2=med.jacobi(A,b)
phi3=med.gauss_seidel(A,b)
phi4=med.sor(A,b)
print(x,y,phi1)
#`\textcolor{darkgreen}{这部分代码可以实现将不同方法生成的图像分别绘制并储存}`
for fn in range(1,5):
	varName='phi%d'%fn
	data=np.c_[x,y,locals()[varName]]
	fig=plt.figure()
	#X,Y=np.meshgrid(np.arange(ba+h,bc,h),np.arange(bd+h,bb,h))
	X,Y=np.meshgrid(np.sort(x),np.sort(y))
	XX=X.flatten()
	YY=Y.flatten()
	A=np.c_[np.ones(data.shape[0]),data[:,:2],np.prod(data[:,:2],axis=1),data[:,:2]**2]
	C,_,_,_=scipy.linalg.lstsq(A,data[:,2])
	Z = np.dot(np.c_[np.ones(XX.shape),XX,YY,XX*YY,XX**2,YY**2],C).reshape(X.shape)
	fig=plt.figure()
	ax=fig.gca(projection='3d')
	surf=ax.plot_surface(X,Y,Z,rstride=1,cstride=1,cmap='rainbow')
	ax.set_xlim(ba,bc)
	ax.set_ylim(bd,bb)
	ax.set_zlim(0,np.max(locals()[varName]))
	ax.zaxis.set_major_locator(LinearLocator(10))
	ax.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
	ax.set_xlabel('x')
	ax.set_ylabel('y')
	ax.set_zlabel('$\phi$')
	fig.colorbar(surf,shrink=0.5,aspect=5)
	fname='figR%d.eps'%fn
	#fname='figC%d.eps'%fn
	plt.savefig(fname)
	plt.clf()
	del data

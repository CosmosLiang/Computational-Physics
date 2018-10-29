import numpy as np
import openpyxl
from openpyxl import Workbook
wb=Workbook()
wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
#`\textcolor{darkgreen}{参数设置}`
ba=-5.0;bb=5.0;bc=5.0;bd=-5.0
nmin=5;nmax=18;x0=2.0;y0=0.0;rho=1
for index in range(nmin,nmax):
	h=(bc-ba)/index;N=(index-1)*(index-1);m=index-1
#`\textcolor{darkgreen}{生成数组矩阵}`
	A=np.zeros(shape=(N,N))
	for i in range(N):
		if(i>=0):
			A[i-1][i]=-1.0
			A[i][i]=4.0
			A[i][i-1]=-1.0
		if(i+m-N<0):
			A[i][(i+m)-N]=-1.0
			A[(i+m)-N][i]=-1.0
	for j in range(0,N,m):
		A[j-1][j]=0;A[j][j-1]=0
	b=np.zeros(shape=(N))
	b[int((x0-ba)/h)+int((y0-bd)/h)]=rho
#`\textcolor{darkgreen}{储存数组矩阵}`'''
	sheet=wb.create_sheet("Sheet%d"%(index-nmin+1))
	for i in range(N):
		sheet.cell(row=i+1,column=N+1,value=b[i])
		for j in range(N):
			sheet.cell(row=i+1,column=j+1,value=A[i][j])
wb.save('Rmatrix.xlsx')
